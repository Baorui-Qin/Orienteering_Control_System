"""运行时逻辑：消费串口消息、驱动 UI 与数据更新、导出成绩。"""

import json
import queue
import tkinter as tk
from datetime import datetime
from tkinter import filedialog, messagebox, simpledialog
from urllib.parse import quote_plus

from openpyxl import Workbook
from serial.tools import list_ports

from calculation import assign_ranks, calculate_result
from .protocol import parse_line


RANKABLE_STATUSES = {"OK", "OT"}
STATUS_COLORS = {
    "OK": "#2b8a3e",
    "OT": "#b8860b",
    "MP": "#409eff",
    "DNF": "#c2410c",
    "DNS": "#86868b",
    "DSQ": "#c92a2a",
}


class RuntimeMixin:
    def _enqueue_line(self, line: str):
        self.msg_q.put(("line", line))
    def _enqueue_state(self, connected: bool, message: str):
        self.msg_q.put(("state", (connected, message)))
    def _process_queue(self):
        """UI 线程消费消息队列。

        串口线程只负责生产消息；所有控件更新都在这里执行，
        避免跨线程直接操作 Tk 组件。
        """
        while True:
            try:
                t, payload = self.msg_q.get_nowait()
            except queue.Empty:
                break

            if t == "state":
                connected, message = payload
                self.conn_state.set(message)
                self.hw_state.set("ONLINE" if connected else "OFFLINE")
                self.hw_card.configure(text_color="#2b8a3e" if connected else "#c92a2a")
            else:
                try:
                    self.handle_line(payload)
                except Exception as exc:  # noqa: BLE001
                    self._append_terminal(f"[ERROR] 处理消息失败: {exc} | {payload}")

        self.after(50, self._process_queue)
    def refresh_ports(self):
        ports = [p.device for p in list_ports.comports()]
        self.port_combo.configure(values=ports if ports else [""])
        if ports and self.port_var.get() not in ports:
            self.port_var.set(ports[0])
        if not ports:
            self.port_var.set("")
    def connect_serial(self):
        port = self.port_var.get().strip()
        if not port:
            messagebox.showwarning("提示", "请先选择 COM 端口")
            return
        try:
            baud = int(self.baud_var.get().strip())
            self.worker.connect(port, baud)
            self._append_terminal(f"[UI] connected {port} @ {baud}")
        except Exception as exc:  # noqa: BLE001
            messagebox.showerror("连接失败", str(exc))
    def disconnect_serial(self):
        self.worker.disconnect()
        self._append_terminal("[UI] disconnected")
    def send_cmd(self, cmd: str):
        try:
            self.worker.send(cmd)
            self._append_terminal(f"[TX] {cmd}")
            return True
        except Exception as exc:  # noqa: BLE001
            messagebox.showwarning("发送失败", str(exc))
            return False
    def send_make_card_cmd(self, command: str, label: str):
        if self.send_cmd(command):
            self.write_state.set(f"{label}模式已下发，请在 5 秒内将卡片贴近 S3")
            self.mode_state.set("WRITE_PENDING")
            self._push_write_history("已下发", f"{label} ({command})")
    def _push_write_history(self, title: str, detail: str = ""):
        ts = datetime.now().strftime("%H:%M:%S")
        line = f"[{ts}] {title}"
        if detail:
            line += f" | {detail}"

        self.write_history_records.append(line)
        self.write_history_records = self.write_history_records[-10:]

        if not hasattr(self, "write_history_box"):
            return

        self.write_history_box.configure(state="normal")
        self.write_history_box.delete("1.0", tk.END)
        for item in reversed(self.write_history_records):
            self.write_history_box.insert(tk.END, item + "\n")
        self.write_history_box.configure(state="disabled")
    def send_custom(self):
        cmd = self.custom_var.get().strip()
        if cmd:
            self.send_cmd(cmd)
            self.custom_var.set("")
    def _send_node_serial_command(self, serial_cmd: str, summary: str):
        if self.send_cmd(serial_cmd):
            self._push_network_feedback("TX", summary)
            return True
        return False
    def send_node_ping(self):
        mac = self._selected_node_mac()
        if not mac:
            messagebox.showwarning("提示", "请先选择一个节点 MAC")
            return
        self._send_node_serial_command(f"CMD:NODE_PING:{mac}", f"PING -> {mac}")
    def send_node_status_request(self):
        mac = self._selected_node_mac()
        if not mac:
            messagebox.showwarning("提示", "请先选择一个节点 MAC")
            return
        self._send_node_serial_command(f"CMD:NODE_GET_STATUS:{mac}", f"GET_STATUS -> {mac}")
    def send_node_sync_time(self):
        mac = self._selected_node_mac()
        if not mac:
            messagebox.showwarning("提示", "请先选择一个节点 MAC")
            return
        self._send_node_serial_command(f"CMD:NODE_SYNC_TIME:{mac}", f"SYNC_TIME -> {mac}")
    def send_nodes_sync_time_all(self):
        macs = []
        seen = set()
        for checkpoint in self.store.list_checkpoints():
            mac = (checkpoint.get("mac") or "").strip().upper()
            if not mac or mac in seen:
                continue
            seen.add(mac)
            macs.append(mac)

        if not macs:
            messagebox.showwarning("提示", "当前没有可群发校时的检查点 MAC")
            return

        joined = ",".join(macs)
        self._send_node_serial_command(f"CMD:NODE_SYNC_TIME_ALL:{joined}", f"SYNC_TIME_ALL -> {len(macs)}")
    def send_node_sync_time_broadcast(self):
        self._send_node_serial_command("CMD:NODE_SYNC_TIME_BROADCAST", "START_SYNC_TIME_BROADCAST -> FFFFFFFFFFFF")
    def stop_node_sync_time_broadcast(self):
        self._send_node_serial_command("CMD:NODE_SYNC_TIME_BROADCAST_STOP", "STOP_SYNC_TIME_BROADCAST")
    def send_node_message(self):
        mac = self._selected_node_mac()
        if not mac:
            messagebox.showwarning("提示", "请先选择一个节点 MAC")
            return
        message = self.node_message_var.get().strip()
        if not message:
            messagebox.showwarning("提示", "请输入要发送的消息")
            return
        encoded = quote_plus(message)
        if self._send_node_serial_command(
            f"CMD:NODE_MESSAGE:{mac}:{encoded}",
            f"MESSAGE -> {mac} | {message}",
        ):
            self.node_message_var.set("")
    def send_wifi_config(self):
        ssid = self.ssid_var.get().strip()
        pwd = self.pwd_var.get().strip()
        if not ssid:
            messagebox.showwarning("提示", "SSID 不能为空")
            return
        if len(ssid) > 32 or len(pwd) > 64:
            messagebox.showwarning("提示", "SSID 最大 32，密码最大 64")
            return
        self.send_cmd(f"CMD:SET_WIFI:{ssid},{pwd}")
    def _status_poller(self):
        if self.worker.is_connected:
            try:
                self.worker.send("CMD:GET_STATUS")
            except Exception:
                pass
        self.after(1000, self._status_poller)
    def handle_line(self, line: str):
        """处理单行协议消息并驱动 UI/数据更新。"""
        self._append_terminal(f"[RX] {line}")
        event = parse_line(line)

        if event.event_type == "sys_msg":
            msg = event.payload.get("msg", "")
            if "退出制卡模式" in msg:
                self.write_state.set("已退出制卡模式，回到网关监听")
                self.mode_state.set("GATEWAY")
            elif any(token in msg for token in ("制卡模式", "检测到卡片", "写入", "空白卡")):
                self.write_state.set(msg)
                self._push_write_history("制卡进度", msg)
            if any(token in msg for token in ("校时", "Wi-Fi")):
                self._push_network_feedback("STATE", msg)
            return

        if event.event_type == "status":
            payload = event.payload
            wifi_ok = payload["wifi"] == 1
            ssid = payload["ssid"] if payload["ssid"] else "-"
            ts = payload["timestamp"]
            dt = datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M:%S") if ts > 0 else "0"

            self.wifi_state.set(f"{'ONLINE' if wifi_ok else 'OFFLINE'} / {ssid}")
            self.ntp_state.set(dt)
            self.last_heartbeat_var.set(datetime.now().strftime("%H:%M:%S"))
            self.wifi_card.configure(text_color="#2b8a3e" if wifi_ok else "#c92a2a")
            self.time_card.configure(text_color="#2b8a3e" if ts > 0 else "#c92a2a")
            self._refresh_device_monitor()
            return

        if event.event_type == "sys_mode":
            self.mode_state.set(event.payload["mode"])
            return

        if event.event_type == "heartbeat":
            hb = event.payload
            self.ntp_state.set(f"{hb.get('time')} ({hb.get('unix')})")
            self.last_heartbeat_var.set(datetime.now().strftime("%H:%M:%S"))
            self.time_card.configure(text_color="#2b8a3e")
            self._refresh_device_monitor()
            return

        if event.event_type == "local_scan":
            uid = event.payload["uid"]
            self._append_scan("LOCAL", uid, self._resolve_runner_name(uid), "SUCCESS")
            return

        if event.event_type == "node_scan":
            mac = (event.payload.get("mac") or event.payload.get("node_id") or "").upper()
            node_id = str(event.payload.get("node_id") or "-")
            last_seen = datetime.now().strftime("%H:%M:%S")
            if mac:
                self.node_last_seen[mac] = last_seen
                state = self.node_status_map.get(mac, {}).copy()
                state.update({"node_id": node_id, "last_seen": last_seen})
                self.node_status_map[mac] = state
            self._append_scan(mac or node_id, event.payload["uid"], self._resolve_runner_name(event.payload["uid"]), "SUCCESS")
            self._refresh_device_monitor()
            self._push_network_feedback("RX", f"SCAN <- {mac or node_id} | UID={event.payload['uid']}")
            return

        if event.event_type == "node_ack":
            payload = event.payload
            mac = payload["mac"]
            state = self.node_status_map.get(mac, {}).copy()
            state.update({"node_id": str(payload["node_id"]), "last_seen": datetime.now().strftime("%H:%M:%S")})
            self.node_status_map[mac] = state
            self.node_last_seen[mac] = state["last_seen"]
            level = "OK" if payload["ok"] else "FAIL"
            detail = payload["msg"] or "-"
            self._push_network_feedback(
                level,
                f"ACK <- {mac} | cmd={payload['cmd']} | seq={payload['seq']} | msg={detail}",
            )
            self._refresh_device_monitor()
            return

        if event.event_type == "node_status":
            payload = event.payload
            mac = payload["mac"]
            status_time = datetime.fromtimestamp(payload["timestamp"]).strftime("%Y-%m-%d %H:%M:%S") if payload["timestamp"] > 0 else "0"
            state = self.node_status_map.get(mac, {}).copy()
            state.update(
                {
                    "node_id": str(payload["node_id"]),
                    "last_seen": datetime.now().strftime("%H:%M:%S"),
                    "ntp": bool(payload["ntp"]),
                    "battery": payload["battery"],
                    "timestamp": payload["timestamp"],
                    "time_text": status_time,
                }
            )
            self.node_status_map[mac] = state
            self.node_last_seen[mac] = state["last_seen"]
            self._push_network_feedback(
                "STATUS",
                f"STATUS <- {mac} | node={payload['node_id']} | ntp={payload['ntp']} | bat={payload['battery']} | time={status_time}",
            )
            self._refresh_device_monitor()
            return

        if event.event_type == "espnow_rx":
            payload = event.payload
            if payload["msg_type"] == "TIME_REQUEST":
                self._push_network_feedback("RX", f"TIME_REQUEST <- {payload['from_mac']}")
                return

            mac = payload["mac"]
            state = self.node_status_map.get(mac, {}).copy()
            state.update(
                {
                    "last_seen": datetime.now().strftime("%H:%M:%S"),
                    "sync_ok": bool(payload["last_sync_ok"]),
                    "battery_mv": payload["battery_mv"],
                    "rtc_timestamp": payload["rtc_timestamp"],
                    "last_sync_timestamp": payload["last_sync_timestamp"],
                    "role": payload["role"],
                }
            )
            self.node_status_map[mac] = state
            self.node_last_seen[mac] = state["last_seen"]
            self._push_network_feedback(
                "RX",
                f"STATUS_REPORT <- {mac} | sync_ok={payload['last_sync_ok']} | bat={payload['battery_mv']}mV",
            )
            self._refresh_device_monitor()
            return

        if event.event_type == "espnow_tx":
            payload = event.payload
            mode = payload["mode"]
            if mode == "MULTICAST":
                self._push_network_feedback("TX", f"TIME_RESPONSE MULTICAST -> {payload.get('count', 0)} | unix={payload['unix']}")
            else:
                self._push_network_feedback(
                    "TX",
                    f"TIME_RESPONSE {mode} -> {payload.get('to_mac', 'FFFFFFFFFFFF')} | unix={payload['unix']}",
                )
            return

        if event.event_type == "espnow_err":
            payload = event.payload
            reason = payload["reason"]
            if reason == "TIME_NOT_SYNCED":
                self._push_network_feedback("ERR", f"{payload['mode']} 时间未同步，请先检查 S3 Wi-Fi/NTP")
            else:
                self._push_network_feedback(
                    "ERR",
                    f"{payload['mode']} <- {payload['from_mac']} | {reason}",
                )
            return

        if event.event_type == "readout":
            self._handle_readout(event.payload["uid"], event.payload["punches"])
            return

        if event.event_type == "write_ok":
            uid = event.payload.get("uid", "-")
            self.write_state.set(f"制卡成功：UID={uid}")
            self.mode_state.set("GATEWAY")
            self._append_terminal(f"[STATE] 制卡成功 UID={uid}")
            self._push_write_history("写卡成功", f"UID={uid}")
            messagebox.showinfo("制卡结果", f"写卡成功\nUID: {uid}")
            return

        if event.event_type == "write_fail":
            uid = event.payload.get("uid")
            self.write_state.set(f"制卡失败{f'：UID={uid}' if uid else ''}")
            self.mode_state.set("GATEWAY")
            self._append_terminal("[STATE] 制卡失败")
            self._push_write_history("写卡失败", f"UID={uid}" if uid else "未返回 UID")
            messagebox.showerror("制卡结果", f"写卡失败{f'\nUID: {uid}' if uid else ''}")
            return

        if event.event_type == "make_success":
            self._append_terminal("[STATE] 制卡成功，自动回到网关模式")
            self._push_write_history("写卡成功", "特殊控制卡")
            messagebox.showinfo("制卡结果", "特殊控制卡写入成功")
            return

        if event.event_type == "make_fail":
            self._append_terminal("[STATE] 制卡失败")
            self._push_write_history("写卡失败", "特殊控制卡")
            messagebox.showerror("制卡结果", "特殊控制卡写入失败")
            return
    def _resolve_runner_name(self, uid: str) -> str:
        runner = self.store.get_runner(uid)
        return runner["name"] if runner and runner.get("name") else "未知"
    def _append_scan(self, node_id: str, uid: str, person: str, status: str):
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.store.insert_scan(ts, node_id, uid, person, status, "")
    def _check_punch_sync(self, punches, checkpoints_map):
        """清算前校验涉及的打卡点是否已同步时间。

        仅在节点状态明确为「未同步」（ntp 为假或 sync_ok 为假）时告警，
        避免在尚未收到节点状态时误报。返回告警文案或空串。
        """
        unsynced = []
        seen = set()
        for mac, _ts in punches or []:
            key = str(mac).upper()
            if key in seen:
                continue
            seen.add(key)
            state = self.node_status_map.get(key)
            if not state:
                continue
            if state.get("ntp") is False or state.get("sync_ok") is False:
                cp_code = (checkpoints_map.get(key) or {}).get("cp_code") or key
                unsynced.append(cp_code)
        if not unsynced:
            return ""
        return "⚠ 时间未同步：" + ", ".join(unsynced) + "（结果时间可能不准）"
    def _handle_readout(self, uid: str, punches):
        runner = self.store.get_runner(uid)
        checkpoints_map = self.store.get_checkpoints_map()
        if not runner:
            self.readout_state.set(f"[DSQ] UID={uid} 未绑定选手")
            self.readout_card.configure(text_color="#c92a2a")
            self._append_terminal(f"[READOUT] 未找到 UID={uid} 的选手映射")
            self._show_readout_popup(
                "未绑定选手",
                [
                    f"UID: {uid}",
                    "成绩有效性: 无效",
                    "原因: 当前卡号还没有绑定选手信息",
                    "",
                    "打卡记录:",
                    *self._build_punch_display_rows(punches, checkpoints_map),
                ],
                "DSQ",
            )
            return

        route_id = runner.get("route_id")
        route = self.store.get_route(route_id) if route_id is not None else None
        if not route:
            person_name = runner.get("name") or "未知"
            self.readout_state.set(f"[DSQ] {runner.get('bib_number', '-') or '-'} {person_name} 未分配路线")
            self.readout_card.configure(text_color="#c92a2a")
            self._append_terminal(f"[READOUT] UID={uid} 缺少 route_id 或路线不存在")
            self._show_readout_popup(
                "路线未配置",
                [
                    f"姓名: {person_name}",
                    f"组别: {runner.get('category') or '-'}",
                    f"号码: {runner.get('bib_number') or '-'}",
                    f"UID: {uid}",
                    "成绩有效性: 无效",
                    "原因: 选手尚未分配路线",
                    "",
                    "打卡记录:",
                    *self._build_punch_display_rows(punches, checkpoints_map),
                ],
                "DSQ",
            )
            return

        route_details = self.store.list_route_details(route["route_id"])
        result = calculate_result(uid, punches, runner, route, route_details, checkpoints_map)
        result["raw_data"] = json.dumps({"punches": punches}, ensure_ascii=False)
        sync_warning = self._check_punch_sync(punches, checkpoints_map)
        result["sync_warning"] = sync_warning or None
        self.store.upsert_result(result)

        person_name = runner.get("name") or "未知"
        bib = runner.get("bib_number") or "-"
        status = result["status"]
        summary = self._format_duration(result["total_seconds"]) if result["total_seconds"] is not None else "--"
        score_text = f" | 分数:{result['final_score']}" if result.get("final_score") is not None else ""
        warn_text = " | ⚠时间未同步" if sync_warning else ""

        self.readout_state.set(f"[{status}] {bib} {person_name} | 用时:{summary}{score_text}{warn_text}")
        if hasattr(self, "readout_banner"):
            self.readout_banner.configure(text_color="#2b8a3e" if status == "OK" else "#409eff" if status == "MP" else "#c92a2a")
        self.readout_card.configure(text_color="#2b8a3e" if status == "OK" else "#409eff" if status == "MP" else "#c92a2a")

        self._append_scan("READOUT", uid, person_name, status)
        self._append_terminal(f"[READOUT] UID={uid} => status={status}, total={summary}, score={result.get('final_score')}")
        self._refresh_leaderboard()

        popup_lines = [
            f"姓名: {person_name}",
            f"组别: {runner.get('category') or '-'}",
            f"号码: {bib}",
            f"UID: {uid}",
            f"路线: {route.get('route_name') or route.get('route_id')}",
            f"模式: {(route.get('race_type') or 'STANDARD').upper()}",
            f"成绩有效性: {self._readout_status_text(status)}",
            f"用时: {summary}",
        ]
        if result.get("final_score") is not None:
            popup_lines.append(f"分数: {result.get('final_score')}")
        if sync_warning:
            popup_lines.append(sync_warning)
        popup_lines.extend(["", "经过检查点:"])
        popup_lines.extend(self._build_punch_display_rows(punches, checkpoints_map))
        self._show_readout_popup(f"{person_name} 刷卡结果", popup_lines, status)
    def _refresh_device_monitor(self):
        if not hasattr(self, "monitor_table"):
            return

        checkpoints = self.store.list_checkpoints()
        ntp_ok = self.ntp_state.get() not in ("TIME UNKNOWN", "0")
        query = self.monitor_search_var.get().strip().lower()

        for item in self.monitor_table.get_children():
            self.monitor_table.delete(item)

        for cp in checkpoints:
            mac = (cp.get("mac") or "").upper()
            node_state = self.node_status_map.get(mac, {})
            role_parts = []
            if cp.get("is_start"):
                role_parts.append("起点")
            if cp.get("is_finish"):
                role_parts.append("终点")
            role = "/".join(role_parts) if role_parts else "普通点"
            node_id = node_state.get("node_id") or "-"
            last_seen = node_state.get("last_seen") or self.node_last_seen.get(mac, "-")
            node_ntp = node_state.get("ntp")
            ntp_text = "SYNCED" if node_ntp else "UNSYNCED" if node_ntp is not None else ("HOST_SYNCED" if ntp_ok else "UNKNOWN")
            if node_state.get("sync_ok") is False:
                ntp_text = "RTC_UNSYNCED"
            battery_text = node_state.get("battery_mv")
            if battery_text is None:
                battery_text = node_state.get("battery")
            battery_text = battery_text if battery_text is not None else "--"
            ready = "YES" if (node_ntp is True and last_seen != "-") else "NO"
            values = (
                cp.get("cp_code") or "-",
                mac,
                node_id,
                role,
                last_seen,
                ntp_text,
                battery_text,
                ready,
            )
            if query and not any(query in str(value).lower() for value in values):
                continue
            self.monitor_table.insert("", tk.END, values=values)
    @staticmethod
    def _format_duration(total_seconds: int):
        minutes = int(total_seconds) // 60
        seconds = int(total_seconds) % 60
        return f"{minutes}分{seconds:02d}秒"
    @staticmethod
    def _punches_from_raw(raw_data):
        """从 results.raw_data 还原 [(mac, ts), ...]。"""
        if not raw_data:
            return []
        try:
            payload = json.loads(raw_data)
        except (TypeError, ValueError, json.JSONDecodeError):
            return []
        out = []
        for p in payload.get("punches") or []:
            if isinstance(p, (list, tuple)) and len(p) == 2:
                out.append((p[0], p[1]))
        return out
    def recalculate_all_results(self):
        """用当前路线/检查点配置重算本赛事全部成绩（保留原始 punches 与裁判改判）。"""
        if not messagebox.askyesno("一键重算", "将用当前路线 / 检查点配置重算本赛事全部成绩，确定继续？"):
            return
        checkpoints_map = self.store.get_checkpoints_map()
        count = 0
        skipped = 0
        for row in self.store.list_results_with_runner():
            uid = row["uid"]
            runner = self.store.get_runner(uid)
            route_id = runner.get("route_id") if runner else None
            route = self.store.get_route(route_id) if route_id is not None else None
            if not runner or not route:
                skipped += 1
                continue
            punches = self._punches_from_raw(row.get("raw_data"))
            route_details = self.store.list_route_details(route["route_id"])
            result = calculate_result(uid, punches, runner, route, route_details, checkpoints_map)
            result["raw_data"] = row.get("raw_data")  # 原始流水不变
            result["sync_warning"] = row.get("sync_warning")
            self.store.upsert_result(result)
            count += 1
        self._refresh_leaderboard()
        self._append_terminal(f"[RECALC] 已重算 {count} 条成绩，跳过 {skipped} 条（缺选手/路线）")
        messagebox.showinfo("一键重算", f"已重算 {count} 条成绩\n跳过 {skipped} 条（缺选手或路线配置）")
    def export_results_excel(self):
        """导出总榜 Excel（按当前排序规则）。"""
        out_path = filedialog.asksaveasfilename(
            title="导出成绩总榜",
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
        )
        if not out_path:
            return

        wb = Workbook()
        ws = wb.active
        rows = sorted(self._merged_results(), key=self._result_sort_key)
        ranks = self._compute_ranks(rows)
        ws.title = "Leaderboard"
        ws.append(["排名", "参赛号", "姓名", "组别", "赛制", "状态", "最终得分", "总用时(秒)", "更新时间"])
        for rank, row in zip(ranks, rows):
            ws.append(
                [
                    rank if rank is not None else "-",
                    row.get("bib_number") or "-",
                    row.get("name") or "-",
                    row.get("category") or "-",
                    row.get("race_type") or "-",
                    row.get("status") or "-",
                    row.get("final_score"),
                    row.get("total_seconds"),
                    row.get("updated_at") or "-",
                ]
            )

        wb.save(out_path)
        messagebox.showinfo("导出完成", f"已导出到\n{out_path}")
    def export_raw_excel(self):
        out_path = filedialog.asksaveasfilename(
            title="导出原始流水",
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
        )
        if not out_path:
            return

        wb = Workbook()
        ws = wb.active
        rows = self.store.list_all()
        ws.title = "Attendance"
        ws.append(["时间戳", "来源节点 ID", "卡号 UID", "绑定人员姓名", "状态", "卡类型"])
        for row in rows:
            ws.append(list(row))

        wb.save(out_path)
        messagebox.showinfo("导出完成", f"已导出到\n{out_path}")
    def export_split_excel(self):
        out_path = filedialog.asksaveasfilename(
            title="导出分段成绩",
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
        )
        if not out_path:
            return

        rows = sorted(self._merged_results(), key=self._result_sort_key)
        ranks = self._compute_ranks(rows)
        checkpoints_map = self.store.get_checkpoints_map()

        wb = Workbook()
        ws = wb.active
        ws.title = "SplitTimes"
        ws.append(["排名", "参赛号", "姓名", "组别", "状态", "总用时(秒)", "分段详情"])

        for rank, row in zip(ranks, rows):
            split_text = self._build_split_text(row.get("raw_data"), checkpoints_map)
            ws.append(
                [
                    rank if rank is not None else "-",
                    row.get("bib_number") or "-",
                    row.get("name") or "-",
                    row.get("category") or "-",
                    row.get("status") or "-",
                    row.get("total_seconds"),
                    split_text,
                ]
            )

        wb.save(out_path)
        messagebox.showinfo("导出完成", f"已导出到\n{out_path}")

    @staticmethod
    def _build_split_text(raw_data: str, checkpoints_map: dict):
        if not raw_data:
            return "-"
        try:
            payload = json.loads(raw_data)
            punches = payload.get("punches") or []
        except (TypeError, ValueError, json.JSONDecodeError):
            return "-"

        normalized = []
        for p in punches:
            if not isinstance(p, (list, tuple)) or len(p) != 2:
                continue
            mac = str(p[0]).upper()
            try:
                ts = int(p[1])
            except (TypeError, ValueError):
                continue
            normalized.append((mac, ts))
        normalized.sort(key=lambda x: x[1])
        if len(normalized) < 2:
            return "-"

        parts = []
        for i in range(1, len(normalized)):
            prev_mac, prev_ts = normalized[i - 1]
            mac, ts = normalized[i]
            delta = max(0, ts - prev_ts)
            prev_code = (checkpoints_map.get(prev_mac) or {}).get("cp_code") or prev_mac[-4:]
            code = (checkpoints_map.get(mac) or {}).get("cp_code") or mac[-4:]
            parts.append(f"{prev_code}->{code}:{delta}s")
        return " | ".join(parts)

    @staticmethod
    def _result_sort_key(row: dict):
        """排行榜排序键。

        - STANDARD: 先 OK 再非 OK，再按总用时升序。
        - SCORE: 先 OK 再非 OK，再按分数降序、用时升序。
        """
        race_type = (row.get("race_type") or "STANDARD").upper()
        status = row.get("status") or "DSQ"
        total = row.get("total_seconds")
        total_key = total if total is not None else 10**12
        bib = row.get("bib_number") or ""

        if race_type == "SCORE":
            score = row.get("final_score")
            score_key = -(score if score is not None else -10**9)
            return (race_type, status != "OK", score_key, total_key, bib)

        return (race_type, status != "OK", total_key, bib)
    @staticmethod
    def _result_rank_key(row: dict):
        """名次并列判定键（去掉 bib 等非成绩字段）。"""
        race_type = (row.get("race_type") or "STANDARD").upper()
        total = row.get("total_seconds")
        total_key = total if total is not None else 10**12
        if race_type == "SCORE":
            score = row.get("final_score")
            score_key = -(score if score is not None else -10**9)
            return (race_type, score_key, total_key)
        return (race_type, total_key)
    @staticmethod
    def _is_rankable(row: dict):
        return (row.get("status") or "").upper() in RANKABLE_STATUSES
    def _compute_ranks(self, sorted_rows: list):
        """对已排序成绩计算并列名次列表（与 sorted_rows 等长）。"""
        return assign_ranks(sorted_rows, key=self._result_rank_key, rankable=self._is_rankable)

    # ------------------------------------------------------------------
    # 成绩合并层（计算结果 ⊕ 裁判改判）
    # ------------------------------------------------------------------
    @staticmethod
    def _opt_int(value):
        if value in (None, ""):
            return None
        try:
            return int(value)
        except (TypeError, ValueError):
            return None
    def _merge_result(self, row: dict) -> dict:
        """把一条成绩与裁判改判覆盖合并为最终展示口径。

        - 状态：有 manual_status 用之，否则用计算状态。
        - 起终点：manual_start_time/manual_finish_time 覆盖后重算用时。
        - 罚分/罚时：penalty_adjust 对 SCORE 加减分、对计时赛加减秒（正=罚时）。
        """
        merged = dict(row)
        race_type = (row.get("race_type") or "STANDARD").upper()

        manual_status = row.get("manual_status")
        if manual_status:
            merged["status"] = manual_status

        has_manual_start = row.get("manual_start_time") not in (None, "")
        has_manual_finish = row.get("manual_finish_time") not in (None, "")
        start_ts = self._opt_int(row.get("manual_start_time"))
        if start_ts is None:
            start_ts = self._opt_int(row.get("start_time"))
        finish_ts = self._opt_int(row.get("manual_finish_time"))
        if finish_ts is None:
            finish_ts = self._opt_int(row.get("finish_time"))

        total = row.get("total_seconds")
        if (has_manual_start or has_manual_finish) and start_ts is not None and finish_ts is not None and finish_ts >= start_ts:
            total = finish_ts - start_ts

        adjust = self._opt_int(row.get("penalty_adjust")) or 0
        if race_type == "SCORE":
            if merged.get("final_score") is not None:
                merged["final_score"] = merged["final_score"] + adjust
        elif total is not None:
            total = total + adjust

        merged["total_seconds"] = total
        merged["overridden"] = bool(
            manual_status or has_manual_start or has_manual_finish or row.get("penalty_adjust") not in (None, "")
        )
        return merged
    def _merged_results(self):
        return [self._merge_result(r) for r in self.store.list_results_with_runner()]

    # ------------------------------------------------------------------
    # 赛事管理
    # ------------------------------------------------------------------
    def _refresh_event_selector(self):
        if not hasattr(self, "event_combo"):
            return
        self.event_option_map = {}
        labels = []
        current_label = None
        for ev in self.store.list_events():
            label = ev["name"] + ("（已归档）" if ev.get("status") == "archived" else "")
            if label in self.event_option_map:
                label = f"{label} #{ev['event_id']}"
            self.event_option_map[label] = ev["event_id"]
            labels.append(label)
            if ev["event_id"] == self.store.current_event_id:
                current_label = label
        self.event_combo.configure(values=labels or [""])
        if current_label is not None:
            self.current_event_var.set(current_label)
        elif labels:
            self.current_event_var.set(labels[0])
    def _reload_all_views(self):
        """切换赛事后刷新所有与赛事相关的视图。"""
        self.refresh_all_setup_views()
        if hasattr(self, "_refresh_org_views"):
            self._refresh_org_views()
        self._refresh_leaderboard()
        self._refresh_device_monitor()
    def _on_event_selected(self, choice: str):
        eid = self.event_option_map.get(choice)
        if eid is None or eid == self.store.current_event_id:
            return
        self.store.set_current_event(eid)
        self._reload_all_views()
        self._append_terminal(f"[EVENT] 切换到赛事「{choice}」")
    def create_event_dialog(self):
        name = simpledialog.askstring("新建赛事", "请输入赛事名称：", parent=self)
        if not name or not name.strip():
            return
        eid = self.store.create_event(name.strip())
        self.store.set_current_event(eid)
        self._refresh_event_selector()
        self._reload_all_views()
        self._append_terminal(f"[EVENT] 新建并切换到赛事「{name.strip()}」")
    def rename_event_dialog(self):
        ev = self.store.get_event(self.store.current_event_id)
        current = ev["name"] if ev else ""
        name = simpledialog.askstring("重命名赛事", "新的赛事名称：", initialvalue=current, parent=self)
        if not name or not name.strip():
            return
        self.store.rename_event(self.store.current_event_id, name.strip(), ev.get("event_date") if ev else None)
        self._refresh_event_selector()
    def archive_current_event(self):
        ev = self.store.get_event(self.store.current_event_id)
        label = ev["name"] if ev else str(self.store.current_event_id)
        if not messagebox.askyesno("归档赛事", f"确认归档赛事「{label}」？归档后仍可切换查看。"):
            return
        self.store.archive_event(self.store.current_event_id)
        self._refresh_event_selector()
    def _append_terminal(self, text: str):
        if not hasattr(self, "terminal"):
            return
        ts = datetime.now().strftime("%H:%M:%S")
        self.terminal.insert(tk.END, f"[{ts}] {text}\n")
        self.terminal.see(tk.END)

    def _sync_device_time(self):
        self.send_node_sync_time()
    def _on_close(self):
        try:
            self.worker.disconnect()
        except Exception:  # noqa: BLE001
            pass
        self.store.close()
        self.destroy()


def main():
    app = S3ControlPanel()
    try:
        app.mainloop()
    except KeyboardInterrupt:
        app._on_close()


if __name__ == "__main__":
    main()
