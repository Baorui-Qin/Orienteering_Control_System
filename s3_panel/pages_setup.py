"""选手页：选手、检查点、路线等基础配置。"""

import re
import tkinter as tk
from datetime import datetime
from tkinter import filedialog, messagebox, ttk

import customtkinter as ctk
from openpyxl import Workbook, load_workbook


class SetupPageMixin:
    def _build_runner_page(self, parent):
        parent.grid_columnconfigure(0, weight=1)
        parent.grid_rowconfigure(0, weight=1)

        runner_frame = ctk.CTkFrame(parent, fg_color="#ffffff", border_width=0, corner_radius=12)
        runner_frame.grid(row=0, column=0, sticky="nsew", padx=12, pady=12)
        self._build_config_page_header(
            runner_frame,
            1,
            "人员绑定",
            "先录入或刷入 UID，再补充参赛号、姓名、组别与路线。",
        )

        form = ctk.CTkFrame(runner_frame, fg_color="transparent")
        form.pack(fill="x", padx=12, pady=6)
        self.runner_uid_entry = self._create_form_row(form, "物理卡号 (UID):", "请刷卡或手输 UID", 0, textvariable=self.runner_uid_var)
        self._create_form_row(form, "参赛号码:", "例如 001", 1, textvariable=self.runner_bib_var)
        self._create_form_row(form, "选手姓名:", "输入真实姓名", 2, textvariable=self.runner_name_var)
        self._create_form_row(form, "所属组别:", "例如 男子精英组", 3, textvariable=self.runner_category_var)
        self._create_form_row(form, "路线编号:", "输入路线 ID", 4, textvariable=self.runner_route_id_var)

        btn_frame = ctk.CTkFrame(form, fg_color="transparent")
        btn_frame.grid(row=0, column=2, rowspan=5, padx=30, sticky="n")
        ctk.CTkButton(btn_frame, text="保存人员", fg_color="#409eff", hover_color="#337ecc", command=self.save_runner).pack(pady=4, fill="x")
        ctk.CTkButton(btn_frame, text="导入名单", fg_color="transparent", hover_color="#66b1ff", text_color="#409eff", command=self.import_runners_from_excel).pack(pady=4, fill="x")
        ctk.CTkButton(btn_frame, text="删除所选", fg_color="#c92a2a", hover_color="#a61e1e", command=self.delete_selected_runner).pack(
            pady=4, fill="x"
        )
        form.grid_columnconfigure(0, weight=0)
        form.grid_columnconfigure(1, weight=1)
        form.grid_columnconfigure(2, weight=0)

        self._build_rule_hint(
            runner_frame,
            "填写规则：UID 必填且唯一；路线编号建议先在步骤 3 创建后再填写；支持保存单条和 Excel 批量导入。",
        )

        self.runner_table = ttk.Treeview(
            runner_frame,
            columns=("uid", "bib", "name", "category", "route"),
            show="headings",
            style="Dark.Treeview",
            height=8,
        )
        self.runner_table.pack(fill="both", expand=True, padx=12, pady=(6, 12))
        self.runner_table.heading("uid", text="UID")
        self.runner_table.heading("bib", text="参赛号")
        self.runner_table.heading("name", text="姓名")
        self.runner_table.heading("category", text="组别")
        self.runner_table.heading("route", text="路线ID")
        self.runner_table.bind("<<TreeviewSelect>>", self._on_runner_select)
    def _build_runners_page(self, parent):
        parent.grid_columnconfigure(0, weight=1)
        parent.grid_rowconfigure(0, weight=1)
        parent.grid_rowconfigure(1, weight=1)

        runner_card, runner_inner = self._make_soft_card(parent, "选手管理", "保留导入、双击编辑和缺失高亮，同时把表格区拆成更清晰的独立数据卡片。")
        runner_card.grid(row=0, column=0, sticky="nsew", padx=8, pady=(0, 12))
        runner_inner.grid_columnconfigure(0, weight=1)
        runner_inner.grid_rowconfigure(1, weight=1)

        runner_form = ctk.CTkFrame(runner_inner, fg_color="transparent", border_width=0)
        runner_form.grid(row=0, column=0, sticky="ew", pady=(14, 14))
        runner_form.grid_columnconfigure((1, 3), weight=1)
        self.runner_uid_entry = self._grid_entry_field(runner_form, 0, 0, "UID", self.runner_uid_var, "刷卡 UID")
        self._grid_entry_field(runner_form, 0, 2, "参赛号", self.runner_bib_var, "例如 101")
        self._grid_entry_field(runner_form, 1, 0, "姓名", self.runner_name_var, "选手姓名")
        self._grid_entry_field(runner_form, 1, 2, "组别", self.runner_category_var, "成人男子")
        self._grid_entry_field(runner_form, 2, 0, "路线 ID", self.runner_route_id_var, "已创建路线编号")
        self._grid_entry_field(runner_form, 2, 2, "发车时间", self.runner_start_var, "可空，YYYY-MM-DD HH:MM:SS")
        ctk.CTkLabel(runner_form, text="单位", text_color="#606266", font=ctk.CTkFont(family="Microsoft YaHei UI", size=12, weight="bold")).grid(row=3, column=0, sticky="w", padx=(0, 8), pady=6)
        self.runner_unit_combo = ctk.CTkComboBox(runner_form, variable=self.runner_unit_var, values=["（无）"], height=36, corner_radius=8, fg_color="#F5F7FA", text_color="#1d1d1f", border_width=1, border_color="#E4E7ED", button_color="#409eff", button_hover_color="#337ecc")
        self.runner_unit_combo.grid(row=3, column=1, sticky="ew", padx=(0, 8), pady=6)
        ctk.CTkLabel(runner_form, text="组别(关联)", text_color="#606266", font=ctk.CTkFont(family="Microsoft YaHei UI", size=12, weight="bold")).grid(row=3, column=2, sticky="w", padx=(12, 8), pady=6)
        self.runner_group_combo = ctk.CTkComboBox(runner_form, variable=self.runner_group_var, values=["（无）"], height=36, corner_radius=8, fg_color="#F5F7FA", text_color="#1d1d1f", border_width=1, border_color="#E4E7ED", button_color="#409eff", button_hover_color="#337ecc")
        self.runner_group_combo.grid(row=3, column=3, sticky="ew", padx=(0, 8), pady=6)

        runner_actions = ctk.CTkFrame(runner_form, fg_color="transparent", border_width=0)
        runner_actions.grid(row=0, column=4, rowspan=4, sticky="ns", padx=(18, 0))
        ctk.CTkButton(runner_actions, text="保存选手", width=110, height=38, corner_radius=14, fg_color="#409eff", hover_color="#337ecc", command=self.save_runner).pack(fill="x", pady=(0, 8))
        ctk.CTkButton(runner_actions, text="导入名单", width=110, height=38, corner_radius=14, fg_color="#ecf5ff", hover_color="#d9ecff", text_color="#409eff", command=self.import_runners_from_excel).pack(fill="x", pady=8)
        ctk.CTkButton(runner_actions, text="清空表单", width=110, height=38, corner_radius=14, fg_color="#f5f5f7", hover_color="#e6f1fc", text_color="#409eff", command=self.clear_runner_form).pack(fill="x", pady=8)

        runner_section, runner_toolbar, runner_table_shell, self.runner_empty_label = self._make_table_section(
            runner_inner,
            "选手列表",
            "列标题固定显示，支持搜索、排序、批量删除和模板导出。",
            "当前还没有选手数据",
        )
        runner_section.grid(row=1, column=0, sticky="nsew")
        self._build_search_bar(
            runner_toolbar,
            self.runner_search_var,
            "搜索 UID / 姓名 / 组别 / 路线",
            self.refresh_runners,
            row=0,
            trailing_buttons=[("删除所选", self.delete_selected_runner, "danger"), ("名单模板", self.export_runner_template, "ghost")],
        )
        self.runner_table = ttk.Treeview(runner_table_shell, columns=("uid", "bib", "name", "category", "route", "start"), show="headings", style="Dark.Treeview", height=9)
        self.runner_table.heading("uid", text="UID")
        self.runner_table.heading("bib", text="参赛号")
        self.runner_table.heading("name", text="姓名")
        self.runner_table.heading("category", text="组别")
        self.runner_table.heading("route", text="路线")
        self.runner_table.heading("start", text="发车时间")
        self.runner_table.column("uid", width=160, anchor="center")
        self.runner_table.column("bib", width=90, anchor="center")
        self.runner_table.column("name", width=130, anchor="w")
        self.runner_table.column("category", width=120, anchor="w")
        self.runner_table.column("route", width=80, anchor="center")
        self.runner_table.column("start", width=150, anchor="center")
        self.runner_table.bind("<<TreeviewSelect>>", self._on_runner_select)
        self.runner_table.bind("<Double-1>", lambda _event: getattr(self, "runner_uid_entry", self.runner_table).focus_set())
        self._mount_treeview(self.runner_table, runner_table_shell, self.runner_empty_label)

        checkpoint_card, checkpoint_inner = self._make_soft_card(parent, "检查点管理", "统一保留表单操作和列表管理，突出点号、MAC 与角色映射关系。")
        checkpoint_card.grid(row=1, column=0, sticky="nsew", padx=8, pady=(0, 8))
        checkpoint_inner.grid_columnconfigure(0, weight=1)
        checkpoint_inner.grid_rowconfigure(1, weight=1)

        checkpoint_form = ctk.CTkFrame(checkpoint_inner, fg_color="transparent", border_width=0)
        checkpoint_form.grid(row=0, column=0, sticky="ew", pady=(14, 14))
        checkpoint_form.grid_columnconfigure(1, weight=1)
        checkpoint_form.grid_columnconfigure(3, weight=1)
        self.cp_mac_entry = self._grid_entry_field(checkpoint_form, 0, 0, "MAC", self.cp_mac_var, "A1B2C3D4E5F6")
        self._grid_entry_field(checkpoint_form, 0, 2, "点号", self.cp_code_var, "31 / CP01")
        toggle_row = ctk.CTkFrame(checkpoint_form, fg_color="transparent", border_width=0)
        toggle_row.grid(row=1, column=0, columnspan=4, sticky="w", pady=(10, 0))
        ctk.CTkCheckBox(toggle_row, text="起点", variable=self.cp_is_start_var, fg_color="#409eff", hover_color="#337ecc", text_color="#1d1d1f").pack(side="left", padx=(0, 12))
        ctk.CTkCheckBox(toggle_row, text="终点", variable=self.cp_is_finish_var, fg_color="#409eff", hover_color="#337ecc", text_color="#1d1d1f").pack(side="left")

        cp_actions = ctk.CTkFrame(checkpoint_form, fg_color="transparent", border_width=0)
        cp_actions.grid(row=0, column=4, rowspan=2, sticky="ns", padx=(18, 0))
        ctk.CTkButton(cp_actions, text="保存检查点", width=110, height=38, corner_radius=14, fg_color="#409eff", hover_color="#337ecc", command=self.save_checkpoint).pack(fill="x", pady=(0, 8))
        ctk.CTkButton(cp_actions, text="新建", width=110, height=38, corner_radius=14, fg_color="#f5f5f7", hover_color="#e6f1fc", text_color="#409eff", command=self.clear_checkpoint_form).pack(fill="x", pady=8)

        checkpoint_section, checkpoint_toolbar, checkpoint_table_shell, self.checkpoint_empty_label = self._make_table_section(
            checkpoint_inner,
            "检查点列表",
            "表头固定显示点号、MAC 与起终点角色，便于现场快速核对。",
            "当前还没有检查点数据",
        )
        checkpoint_section.grid(row=1, column=0, sticky="nsew")
        self._build_search_bar(
            checkpoint_toolbar,
            self.checkpoint_search_var,
            "搜索点号 / MAC / 角色",
            self.refresh_checkpoints,
            row=0,
            trailing_buttons=[("删除所选", self.delete_selected_checkpoint, "danger"), ("点位模板", self.export_checkpoint_template, "ghost")],
        )
        self.cp_table = ttk.Treeview(checkpoint_table_shell, columns=("mac", "code", "start", "finish"), show="headings", style="Dark.Treeview", height=8)
        self.cp_table.heading("mac", text="MAC")
        self.cp_table.heading("code", text="点号")
        self.cp_table.heading("start", text="起点")
        self.cp_table.heading("finish", text="终点")
        self.cp_table.column("mac", width=180, anchor="center")
        self.cp_table.column("code", width=120, anchor="center")
        self.cp_table.column("start", width=90, anchor="center")
        self.cp_table.column("finish", width=90, anchor="center")
        self.cp_table.bind("<<TreeviewSelect>>", self._on_checkpoint_select)
        self.cp_table.bind("<Double-1>", lambda _event: getattr(self, "cp_mac_entry", self.cp_table).focus_set())
        self._mount_treeview(self.cp_table, checkpoint_table_shell, self.checkpoint_empty_label)
    def _build_checkpoint_page(self, parent):
        parent.grid_columnconfigure(0, weight=1)
        parent.grid_rowconfigure(0, weight=1)

        cp_frame = ctk.CTkFrame(parent, fg_color="#ffffff", border_width=0, corner_radius=12)
        cp_frame.grid(row=0, column=0, sticky="nsew", padx=12, pady=12)
        self._build_config_page_header(
            cp_frame,
            2,
            "检查点映射",
            "录入设备 MAC 与点号，并标记起点/终点角色。",
        )

        cp_form = ctk.CTkFrame(cp_frame, fg_color="transparent")
        cp_form.pack(fill="x", padx=12, pady=6)
        self.cp_mac_entry = self._create_form_row(cp_form, "设备 MAC:", "输入 12 位十六进制 MAC", 0, textvariable=self.cp_mac_var)
        self._create_form_row(cp_form, "检查点编号:", "例如 CP01 / 31", 1, textvariable=self.cp_code_var)

        flag_row = ctk.CTkFrame(cp_form, fg_color="transparent")
        flag_row.grid(row=2, column=1, sticky="w", padx=(5, 10), pady=5)
        ctk.CTkCheckBox(flag_row, text="起点", variable=self.cp_is_start_var).pack(side="left", padx=(0, 14))
        ctk.CTkCheckBox(flag_row, text="终点", variable=self.cp_is_finish_var).pack(side="left")

        btn_frame = ctk.CTkFrame(cp_form, fg_color="transparent")
        btn_frame.grid(row=0, column=2, rowspan=3, padx=30, sticky="n")
        ctk.CTkButton(
            btn_frame,
            text="新建",
            fg_color="transparent",
            hover_color="#f5f5f7",
            text_color="#409eff",
            command=self.clear_checkpoint_form,
        ).pack(pady=4, fill="x")
        ctk.CTkButton(btn_frame, text="保存检查点", fg_color="#409eff", hover_color="#337ecc", command=self.save_checkpoint).pack(pady=4, fill="x")
        ctk.CTkButton(btn_frame, text="删除所选", fg_color="#c92a2a", hover_color="#a61e1e", command=self.delete_selected_checkpoint).pack(
            pady=4, fill="x"
        )

        cp_form.grid_columnconfigure(0, weight=0)
        cp_form.grid_columnconfigure(1, weight=1)
        cp_form.grid_columnconfigure(2, weight=0)

        self._build_rule_hint(
            cp_frame,
            "填写规则：MAC 必须是 12 位十六进制（如 A1B2C3D4E5F6）；同一 MAC 只能对应一个检查点；起点和终点可按赛道规则勾选。",
        )

        self.cp_table = ttk.Treeview(
            cp_frame,
            columns=("mac", "code", "start", "finish"),
            show="headings",
            style="Dark.Treeview",
            height=8,
        )
        self.cp_table.pack(fill="both", expand=True, padx=12, pady=(6, 12))
        self.cp_table.heading("mac", text="MAC")
        self.cp_table.heading("code", text="点号")
        self.cp_table.heading("start", text="起点")
        self.cp_table.heading("finish", text="终点")
        self.cp_table.bind("<<TreeviewSelect>>", self._on_checkpoint_select)
    @staticmethod
    def _parse_start_time(text: str):
        """把发车时间文本解析为 unix 时间戳；空串返回 None，非法抛 ValueError。"""
        text = (text or "").strip()
        if not text:
            return None
        if text.isdigit():
            return int(text)
        dt = datetime.strptime(text, "%Y-%m-%d %H:%M:%S")
        return int(dt.timestamp())

    @staticmethod
    def _format_start_time(ts):
        if ts in (None, ""):
            return ""
        try:
            return datetime.fromtimestamp(int(ts)).strftime("%Y-%m-%d %H:%M:%S")
        except (TypeError, ValueError, OSError):
            return ""

    def save_runner(self):
        uid = self.runner_uid_var.get().strip()
        if not uid:
            messagebox.showwarning("提示", "UID 不能为空")
            return
        route_text = self.runner_route_id_var.get().strip()
        try:
            route_id = int(route_text) if route_text else None
        except ValueError:
            messagebox.showwarning("提示", "路线ID 必须是整数")
            return
        try:
            start_time = self._parse_start_time(self.runner_start_var.get())
        except ValueError:
            messagebox.showwarning("提示", "发车时间格式应为 YYYY-MM-DD HH:MM:SS 或留空")
            return

        unit_id = self.runner_unit_option_map.get(self.runner_unit_var.get())
        group_id = self.runner_group_option_map.get(self.runner_group_var.get())
        try:
            self.store.upsert_runner(
                uid=uid,
                bib_number=self.runner_bib_var.get().strip(),
                name=self.runner_name_var.get().strip(),
                category=self.runner_category_var.get().strip(),
                route_id=route_id,
                start_time=start_time,
                unit_id=unit_id,
                group_id=group_id,
            )
        except sqlite3.Error as exc:
            messagebox.showerror("保存失败", str(exc))
            return
        self.refresh_runners()
        self._refresh_leaderboard()
    def import_runners_from_excel(self):
        file_path = filedialog.askopenfilename(
            title="导入名单",
            filetypes=[("Excel Files", "*.xlsx;*.xlsm")],
        )
        if not file_path:
            return

        try:
            wb = load_workbook(file_path)
            ws = wb.active
        except Exception as exc:  # noqa: BLE001
            messagebox.showerror("读取失败", str(exc))
            return

        ok_count = 0
        for row_idx, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            uid = str(values[0]).strip() if values and values[0] is not None else ""
            if not uid:
                continue
            bib = str(values[1]).strip() if len(values) > 1 and values[1] is not None else ""
            name = str(values[2]).strip() if len(values) > 2 and values[2] is not None else ""
            category = str(values[3]).strip() if len(values) > 3 and values[3] is not None else ""
            route_id = None
            if len(values) > 4 and values[4] is not None and str(values[4]).strip():
                try:
                    route_id = int(values[4])
                except ValueError:
                    self._append_terminal(f"[IMPORT] 第{row_idx}行 route_id 非整数，已跳过该字段")
            start_time = None
            if len(values) > 5 and values[5] is not None and str(values[5]).strip():
                try:
                    start_time = self._parse_start_time(str(values[5]).strip())
                except ValueError:
                    self._append_terminal(f"[IMPORT] 第{row_idx}行 发车时间格式非法，已跳过该字段")

            try:
                self.store.upsert_runner(uid=uid, bib_number=bib, name=name, category=category, route_id=route_id, start_time=start_time)
                ok_count += 1
            except sqlite3.Error as exc:
                self._append_terminal(f"[IMPORT] 第{row_idx}行写入失败: {exc}")

        self.refresh_runners()
        self._refresh_leaderboard()
        messagebox.showinfo("导入完成", f"成功导入/更新 {ok_count} 条选手记录")
    def delete_selected_runner(self):
        selected = self.runner_table.selection() if hasattr(self, "runner_table") else ()
        if not selected:
            messagebox.showwarning("提示", "请先选择一条人员记录")
            return
        item = self.runner_table.item(selected[0])
        uid = (item.get("values") or [""])[0]
        if not uid:
            return
        try:
            self.store.delete_runner(str(uid))
        except sqlite3.Error as exc:
            messagebox.showerror("删除失败", str(exc))
            return
        self.refresh_runners()
        self._refresh_leaderboard()
    def delete_selected_checkpoint(self):
        selected = self.cp_table.selection() if hasattr(self, "cp_table") else ()
        if not selected:
            messagebox.showwarning("提示", "请先选择一个检查点")
            return
        item = self.cp_table.item(selected[0])
        mac = (item.get("values") or [""])[0]
        if not mac:
            return
        try:
            self.store.delete_checkpoint(str(mac))
        except sqlite3.Error as exc:
            messagebox.showerror("删除失败", str(exc))
            return
        self.refresh_checkpoints()
        self._set_detail_checkpoint(mac)
        if hasattr(self, "cp_table"):
            self._select_tree_item_by_value(self.cp_table, mac)
    def _on_runner_select(self, _event=None):
        selected = self.runner_table.selection() if hasattr(self, "runner_table") else ()
        if not selected:
            return
        values = self.runner_table.item(selected[0]).get("values") or []
        if len(values) < 5:
            return
        self.runner_uid_var.set(str(values[0]))
        self.runner_bib_var.set(str(values[1]))
        self.runner_name_var.set(str(values[2]))
        self.runner_category_var.set(str(values[3]))
        self.runner_route_id_var.set(str(values[4]))
        self.runner_start_var.set(str(values[5]) if len(values) > 5 else "")
        runner = self.store.get_runner(str(values[0]))
        if runner:
            self.runner_unit_var.set(self._label_for_id(self.runner_unit_option_map, runner.get("unit_id")))
            self.runner_group_var.set(self._label_for_id(self.runner_group_option_map, runner.get("group_id")))
    def _on_checkpoint_select(self, _event=None):
        selected = self.cp_table.selection() if hasattr(self, "cp_table") else ()
        if not selected:
            return
        values = self.cp_table.item(selected[0]).get("values") or []
        if len(values) < 4:
            return
        self.cp_mac_var.set(str(values[0]))
        self.cp_code_var.set(str(values[1]))
        self.cp_is_start_var.set(str(values[2]).upper() == "Y")
        self.cp_is_finish_var.set(str(values[3]).upper() == "Y")
        self._set_detail_checkpoint(str(values[0]))
    def clear_checkpoint_form(self):
        self.cp_mac_var.set("")
        self.cp_code_var.set("")
        self.cp_is_start_var.set(False)
        self.cp_is_finish_var.set(False)
        if hasattr(self, "cp_table"):
            self.cp_table.selection_remove(self.cp_table.selection())
    def save_checkpoint(self):
        mac = self.cp_mac_var.get().strip().upper()
        if not mac or not re.fullmatch(r"[0-9A-F]{12}", mac):
            messagebox.showwarning("提示", "MAC 需为 12 位十六进制")
            return
        try:
            self.store.upsert_checkpoint(
                mac=mac,
                cp_code=self.cp_code_var.get().strip(),
                is_start=self.cp_is_start_var.get(),
                is_finish=self.cp_is_finish_var.get(),
            )
        except sqlite3.Error as exc:
            messagebox.showerror("保存失败", str(exc))
            return
        self.refresh_checkpoints()
        self._set_detail_checkpoint(mac)
        if hasattr(self, "cp_table"):
            self._select_tree_item_by_value(self.cp_table, mac)
    def refresh_all_setup_views(self):
        self.refresh_runners()
        self.refresh_checkpoints()
        self.refresh_routes()
        self.refresh_route_details()
    def clear_runner_form(self):
        self.runner_uid_var.set("")
        self.runner_bib_var.set("")
        self.runner_name_var.set("")
        self.runner_category_var.set("")
        self.runner_route_id_var.set("")
        self.runner_start_var.set("")
        self.runner_unit_var.set("（无）")
        self.runner_group_var.set("（无）")
        if hasattr(self, "runner_table"):
            self.runner_table.selection_remove(self.runner_table.selection())
    def refresh_runners(self):
        if not hasattr(self, "runner_table"):
            return
        current_uid = self.runner_uid_var.get().strip()
        query = self.runner_search_var.get().strip().lower()
        for item in self.runner_table.get_children():
            self.runner_table.delete(item)
        for row in self.store.list_runners():
            values = (
                row.get("uid") or "",
                row.get("bib_number") or "",
                row.get("name") or "",
                row.get("category") or "",
                row.get("route_id") if row.get("route_id") is not None else "",
                self._format_start_time(row.get("start_time")),
            )
            if not self._row_matches_query(list(values), query):
                continue
            tags = ("missing",) if (not values[1] or not values[2] or values[4] == "") else ()
            self.runner_table.insert("", tk.END, values=values, tags=tags)
        if current_uid:
            self._select_tree_item_by_value(self.runner_table, current_uid)
        self._reapply_tree_sort(self.runner_table)
    def refresh_checkpoints(self):
        if not hasattr(self, "cp_table"):
            return
        current_mac = self.cp_mac_var.get().strip().upper()
        query = self.checkpoint_search_var.get().strip().lower()
        for item in self.cp_table.get_children():
            self.cp_table.delete(item)
        for row in self.store.list_checkpoints():
            role_text = self._checkpoint_role_text(row)
            values = (
                row.get("mac") or "",
                row.get("cp_code") or "",
                "Y" if row.get("is_start") else "N",
                "Y" if row.get("is_finish") else "N",
            )
            if not self._row_matches_query([values[0], values[1], role_text], query):
                continue
            tags = ("missing",) if not values[1] else ()
            self.cp_table.insert("", tk.END, values=values, tags=tags)
        self._refresh_detail_checkpoint_options()
        self._refresh_available_checkpoint_table()
        if current_mac:
            self._select_tree_item_by_value(self.cp_table, current_mac)
        self._reapply_tree_sort(self.cp_table)
        self._refresh_device_monitor()
    def export_runner_template(self):
        self._export_template("选手名单模板", "RunnerTemplate", ["UID", "BibNumber", "Name", "Category", "RouteID", "StartTime(YYYY-MM-DD HH:MM:SS)"])
    def export_checkpoint_template(self):
        self._export_template("检查点模板", "CheckpointTemplate", ["MAC", "CPCode", "IsStart(Y/N)", "IsFinish(Y/N)"])
    def _export_template(self, dialog_title: str, sheet_name: str, headers: list):
        out_path = filedialog.asksaveasfilename(
            title=dialog_title,
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
        )
        if not out_path:
            return
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(headers)
        wb.save(out_path)
        messagebox.showinfo("模板已导出", f"已保存到\n{out_path}")
